import openpyxl
from openpyxl import Workbook,load_workbook

wb = load_workbook('test.xlsx')

sheet = wb.active

for row in sheet.iter_rows(min_row=4, max_row=20,max_col=5,values_only=True):
    print(row)



